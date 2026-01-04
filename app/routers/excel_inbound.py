from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from openpyxl import load_workbook

from app.db import upsert_inventory, add_history
from app.utils.excel_kor_columns import build_col_index, validate_required

router = APIRouter(prefix="/api/excel", tags=["excel-inbound"])

def _norm(v):
    if v is None:
        return ""
    return str(v).strip()

@router.post("/inbound")
async def inbound_excel(
    file: UploadFile = File(...),
    default_warehouse: str = Form("기본"),
    worker: str = Form(""),
):
    """입고 엑셀 업로드: 재고 +, 이력 INBOUND_EXCEL 기록"""
    if not file.filename.lower().endswith((".xlsx",".xlsm",".xltx",".xltm")):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드 가능합니다.")

    wb = load_workbook(file.file, data_only=True)
    ws = wb.active

    headers = [_norm(c.value) for c in ws[1]]
    idx = build_col_index(headers)
    ok, missing = validate_required(idx)
    if not ok:
        raise HTTPException(status_code=400, detail=f"필수 컬럼 누락: {', '.join(missing)}")

    applied = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue

        location = _norm(row[idx["로케이션"]])
        item_code = _norm(row[idx["품번"]])
        item_name = _norm(row[idx["품명"]])
        lot = _norm(row[idx["LOT"]])
        spec = _norm(row[idx["규격"]])
        qty_raw = row[idx["수량"]]

        # 완전 빈 줄 skip
        if location=="" and item_code=="" and item_name=="" and lot=="" and spec=="" and (qty_raw is None or str(qty_raw).strip()==""):
            skipped += 1
            continue

        try:
            qty = int(float(qty_raw))
        except Exception:
            raise HTTPException(status_code=400, detail=f"수량 숫자 변환 실패: {qty_raw}")

        if qty <= 0:
            raise HTTPException(status_code=400, detail=f"수량은 1 이상이어야 합니다. (문제 row={applied+skipped+2})")

        warehouse = default_warehouse
        if "창고" in idx:
            wh = _norm(row[idx["창고"]])
            if wh:
                warehouse = wh

        note = ""
        if "비고" in idx:
            note = _norm(row[idx["비고"]])

        # 재고 반영
        upsert_inventory(
            warehouse=warehouse,
            location=location,
            item_code=item_code,
            item_name=item_name,
            lot=lot,
            spec=spec,
            qty_delta=qty,
            note=note,
        )

        hist_note = note
        if worker:
            hist_note = f"[작업자:{worker}] " + (note or "")

        add_history(
            type_="INBOUND_EXCEL",
            warehouse=warehouse,
            item_code=item_code,
            item_name=item_name,
            lot=lot,
            spec=spec,
            from_location="",
            to_location=location,
            qty=qty,
            note=hist_note,
        )

        applied += 1

    return {"ok": True, "applied": applied, "skipped": skipped}
