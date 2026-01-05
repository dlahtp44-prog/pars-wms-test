from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from openpyxl import load_workbook
from datetime import datetime
from app.db import db
router=APIRouter(prefix="/api/excel", tags=["excel"])
REQ_IN=["로케이션","품번","품명","LOT","규격","수량"]
REQ_OUT=["로케이션","품번","LOT","규격","수량"]
REQ_MOVE=["출발","도착","품번","LOT","규격","수량"]
def _read(file:UploadFile):
    wb=load_workbook(file.file, data_only=True); ws=wb.active
    headers=[str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx={h:i for i,h in enumerate(headers)}
    return ws, idx
def _get(ws,r,idx,k):
    i=idx.get(k); 
    if i is None: return ""
    v=ws.cell(row=r,column=i+1).value
    return "" if v is None else str(v).strip()
@router.post("/inbound")
def inbound_upload(warehouse:str=Form("MAIN"), file:UploadFile=File(...)):
    if not file.filename.lower().endswith((".xlsx",".xlsm")): raise HTTPException(400,"xlsx만 가능")
    ws,idx=_read(file)
    for k in REQ_IN:
        if k not in idx: raise HTTPException(400,f"필수 컬럼 누락: {k}")
    now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"); count=0
    with db() as conn:
        cur=conn.cursor()
        for r in range(2, ws.max_row+1):
            loc=_get(ws,r,idx,"로케이션"); code=_get(ws,r,idx,"품번"); name=_get(ws,r,idx,"품명")
            lot=_get(ws,r,idx,"LOT"); spec=_get(ws,r,idx,"규격"); qtys=_get(ws,r,idx,"수량")
            brand=_get(ws,r,idx,"브랜드"); note=_get(ws,r,idx,"비고")
            if not (loc and code and lot and spec and qtys): continue
            try: qty=int(float(qtys))
            except: continue
            if qty<=0: continue
            row=cur.execute("SELECT id,qty FROM inventory WHERE warehouse=? AND location=? AND item_code=? AND lot=? AND spec=?",
                            (warehouse,loc,code,lot,spec)).fetchone()
            if row:
                cur.execute("UPDATE inventory SET qty=?, item_name=?, brand=?, note=?, updated_at=? WHERE id=?",
                            (int(row['qty'])+qty,name,brand,note,now,row['id']))
            else:
                cur.execute("INSERT INTO inventory(warehouse,location,brand,item_code,item_name,lot,spec,qty,note,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
                            (warehouse,loc,brand,code,name,lot,spec,qty,note,now))
            cur.execute("INSERT INTO history(created_at,type,warehouse,location,brand,item_code,item_name,lot,spec,qty,note,operator) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                        (now,"INBOUND",warehouse,loc,brand,code,name,lot,spec,qty,note,"excel"))
            count+=1
    return {"ok":True,"rows":count}
@router.post("/outbound")
def outbound_upload(warehouse:str=Form("MAIN"), file:UploadFile=File(...)):
    if not file.filename.lower().endswith((".xlsx",".xlsm")): raise HTTPException(400,"xlsx만 가능")
    ws,idx=_read(file)
    for k in REQ_OUT:
        if k not in idx: raise HTTPException(400,f"필수 컬럼 누락: {k}")
    now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"); count=0
    with db() as conn:
        cur=conn.cursor()
        for r in range(2, ws.max_row+1):
            loc=_get(ws,r,idx,"로케이션"); code=_get(ws,r,idx,"품번"); lot=_get(ws,r,idx,"LOT"); spec=_get(ws,r,idx,"규격"); qtys=_get(ws,r,idx,"수량")
            note=_get(ws,r,idx,"비고")
            if not (loc and code and lot and spec and qtys): continue
            try: qty=int(float(qtys))
            except: continue
            if qty<=0: continue
            inv=cur.execute("SELECT id,qty,item_name,brand FROM inventory WHERE warehouse=? AND location=? AND item_code=? AND lot=? AND spec=?",
                            (warehouse,loc,code,lot,spec)).fetchone()
            if not inv or int(inv["qty"])<qty: continue
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?",(int(inv["qty"])-qty,now,inv["id"]))
            cur.execute("INSERT INTO history(created_at,type,warehouse,location,brand,item_code,item_name,lot,spec,qty,note,operator) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                        (now,"OUTBOUND",warehouse,loc,inv["brand"],code,inv["item_name"],lot,spec,qty,note,"excel"))
            count+=1
    return {"ok":True,"rows":count}
@router.post("/move")
def move_upload(warehouse:str=Form("MAIN"), file:UploadFile=File(...)):
    if not file.filename.lower().endswith((".xlsx",".xlsm")): raise HTTPException(400,"xlsx만 가능")
    ws,idx=_read(file)
    for k in REQ_MOVE:
        if k not in idx: raise HTTPException(400,f"필수 컬럼 누락: {k}")
    now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"); count=0
    with db() as conn:
        cur=conn.cursor()
        for r in range(2, ws.max_row+1):
            frm=_get(ws,r,idx,"출발"); to=_get(ws,r,idx,"도착"); code=_get(ws,r,idx,"품번"); lot=_get(ws,r,idx,"LOT"); spec=_get(ws,r,idx,"규격"); qtys=_get(ws,r,idx,"수량")
            note=_get(ws,r,idx,"비고")
            if not (frm and to and code and lot and spec and qtys): continue
            try: qty=int(float(qtys))
            except: continue
            if qty<=0: continue
            src=cur.execute("SELECT id,qty,item_name,brand FROM inventory WHERE warehouse=? AND location=? AND item_code=? AND lot=? AND spec=?",
                            (warehouse,frm,code,lot,spec)).fetchone()
            if not src or int(src["qty"])<qty: continue
            cur.execute("UPDATE inventory SET qty=?, updated_at=? WHERE id=?",(int(src["qty"])-qty,now,src["id"]))
            dst=cur.execute("SELECT id,qty FROM inventory WHERE warehouse=? AND location=? AND item_code=? AND lot=? AND spec=?",
                            (warehouse,to,code,lot,spec)).fetchone()
            if dst:
                cur.execute("UPDATE inventory SET qty=?, item_name=?, brand=?, updated_at=? WHERE id=?",
                            (int(dst["qty"])+qty,src["item_name"],src["brand"],now,dst["id"]))
            else:
                cur.execute("INSERT INTO inventory(warehouse,location,brand,item_code,item_name,lot,spec,qty,note,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
                            (warehouse,to,src["brand"],code,src["item_name"],lot,spec,qty,note,now))
            cur.execute("INSERT INTO history(created_at,type,warehouse,from_location,to_location,brand,item_code,item_name,lot,spec,qty,note,operator) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (now,"MOVE",warehouse,frm,to,src["brand"],code,src["item_name"],lot,spec,qty,note,"excel"))
            count+=1
    return {"ok":True,"rows":count}
